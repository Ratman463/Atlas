"""
Embedding module for document processing and vector generation.
"""
import os
from typing import List, Tuple
from sentence_transformers import SentenceTransformer
from langchain_core.embeddings import Embeddings
from langchain_experimental.text_splitter import SemanticChunker
import config


class SentenceTransformerEmbeddings(Embeddings):
    """Wrapper around SentenceTransformer to implement LangChain Embeddings interface.
    
    This avoids the TypeError('Protocols cannot be instantiated') that occurs
    when using HuggingFaceEmbeddings from langchain_community.
    """

    def __init__(self, model: SentenceTransformer):
        self.model = model

    def embed_documents(self, texts: List[str]) -> List[List[float]]:
        """Embed a list of documents."""
        embeddings = self.model.encode(texts, convert_to_numpy=True)
        return [emb.tolist() for emb in embeddings]

    def embed_query(self, text: str) -> List[float]:
        """Embed a single query text."""
        embedding = self.model.encode(text, convert_to_numpy=True)
        return embedding.tolist()


class EmbeddingEngine:
    """Handles text embedding and document processing."""
    
    def __init__(self):
        """Initialize the embedding model and semantic chunker."""
        self.model = SentenceTransformer(config.EMBEDDING_MODEL)
        # Use a LangChain-compatible embeddings wrapper to avoid
        # the HuggingFaceEmbeddings Protocol instantiation bug
        embeddings = SentenceTransformerEmbeddings(self.model)
        self.text_splitter = SemanticChunker(
            embeddings,
            breakpoint_threshold_type="percentile"
        )

    def semantic_chunk_text(self, text: str) -> List[str]:
        """
        Split text into semantic chunks using SemanticChunker.
        
        Leverages LangChain's SemanticChunker which uses cosine similarity
        between consecutive sentences to detect topic boundaries.
        
        Args:
            text: Text to split
            
        Returns:
            List of text chunks
        """
        if not text.strip():
            return []
        documents = self.text_splitter.create_documents([text])
        return [doc.page_content for doc in documents]
    
    def chunk_text(self, text: str, chunk_size: int = config.CHUNK_SIZE, 
                   overlap: int = config.CHUNK_OVERLAP) -> List[str]:
        """
        Split text into overlapping chunks.
        
        Args:
            text: Text to split
            chunk_size: Maximum characters per chunk
            overlap: Number of overlapping characters between chunks
            
        Returns:
            List of text chunks
        """
        if len(text) <= chunk_size:
            return [text.strip()] if text.strip() else []
        
        chunks = []
        start = 0
        
        while start < len(text):
            end = start + chunk_size
            
            # Try to find a good break point (sentence end)
            if end < len(text):
                # Look for sentence endings
                for break_char in ['.', '!', '?', '\n']:
                    last_break = text.rfind(break_char, start, end)
                    if last_break > start:
                        end = last_break + 1
                        break
            
            chunk = text[start:end].strip()
            if chunk:
                chunks.append(chunk)
            
            # Move start with overlap
            start = end - overlap if end < len(text) else end
        
        return chunks
    
    def embed_text(self, text: str) -> List[float]:
        """
        Generate embedding for a text.
        
        Args:
            text: Text to embed
            
        Returns:
            Vector embedding as list
        """
        embedding = self.model.encode(text, convert_to_numpy=True)
        return embedding.tolist()
    
    def embed_chunks(self, chunks: List[str]) -> List[List[float]]:
        """
        Generate embeddings for multiple chunks.
        
        Args:
            chunks: List of text chunks
            
        Returns:
            List of vector embeddings
        """
        embeddings = self.model.encode(chunks, convert_to_numpy=True)
        return [emb.tolist() for emb in embeddings]
    
    def process_text(self, text: str) -> List[Tuple[str, List[float]]]:
        """
        Process text: chunk and embed.
        
        Args:
            text: Full text content
            
        Returns:
            List of (chunk, embedding) tuples
        """
        chunks = self.semantic_chunk_text(text)
        if not chunks:
            return []
        
        embeddings = self.embed_chunks(chunks)
        return list(zip(chunks, embeddings))


class DocumentProcessor:
    """Handles document file processing."""
    
    @staticmethod
    def read_txt(file_path: str) -> str:
        """Read plain text file (strip UTF-8 BOM)."""
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            return f.read()

    @classmethod
    def read_markdown(cls, file_path: str) -> str:
        """Read Markdown — treat as plain text (the chunker handles prose fine)."""
        return cls.read_txt(file_path)
    
    @staticmethod
    def read_pdf(file_path: str) -> str:
        """Read PDF file."""
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        text_parts = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)
        return '\n'.join(text_parts)
    
    @staticmethod
    def read_docx(file_path: str) -> str:
        """Read DOCX file."""
        from docx import Document
        doc = Document(file_path)
        text_parts = [paragraph.text for paragraph in doc.paragraphs if paragraph.text]
        return '\n'.join(text_parts)

    @staticmethod
    def read_excel(file_path: str) -> str:
        """Read Excel (.xlsx / .xls) — extract all non-empty cells as text rows.

        Each row becomes a line like: [Sheet1] A1: value | B2: value ...
        """
        ext = os.path.splitext(file_path)[1].lower()
        lines: list[str] = []

        def _cell_str(v) -> str:
            if v is None:
                return ""
            if isinstance(v, float) and v.is_integer():
                return str(int(v))
            return str(v).strip()

        if ext == '.xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                rows = []
                for row in ws.iter_rows():
                    vals = [_cell_str(c.value) for c in row]
                    # 整行全空则跳过
                    if not any(vals):
                        continue
                    # 去尾部空列
                    while vals and vals[-1] == "":
                        vals.pop()
                    rows.append(" | ".join(vals))
                if rows:
                    lines.append(f"[Sheet: {ws.title}]")
                    lines.extend(rows)
            wb.close()
        else:  # .xls
            import xlrd
            book = xlrd.open_workbook(file_path)
            for sheet in book.sheets():
                rows = []
                for r in range(sheet.nrows):
                    vals = [_cell_str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
                    if not any(vals):
                        continue
                    while vals and vals[-1] == "":
                        vals.pop()
                    rows.append(" | ".join(vals))
                if rows:
                    lines.append(f"[Sheet: {sheet.name}]")
                    lines.extend(rows)

        return "\n".join(lines)

    @classmethod
    def read_document(cls, file_path: str) -> str:
        """
        Read document based on extension.
        
        Args:
            file_path: Path to the document
            
        Returns:
            Extracted text content
        """
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext == '.txt':
            return cls.read_txt(file_path)
        elif ext in ('.md', '.markdown'):
            return cls.read_markdown(file_path)
        elif ext == '.pdf':
            return cls.read_pdf(file_path)
        elif ext == '.docx':
            return cls.read_docx(file_path)
        elif ext in ('.xlsx', '.xls'):
            return cls.read_excel(file_path)
        else:
            raise ValueError(f"Unsupported file extension: {ext}")


# Global instances
_embedding_engine: EmbeddingEngine = None


def get_embedding_engine() -> EmbeddingEngine:
    """Get or create embedding engine instance."""
    global _embedding_engine
    if _embedding_engine is None:
        _embedding_engine = EmbeddingEngine()
    return _embedding_engine